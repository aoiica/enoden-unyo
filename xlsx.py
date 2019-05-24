import openpyxl
import pprint



wb = openpyxl.load_workbook('./pyunyo.xlsx')

sheet = wb['pypy - teire']
'''
cell = sheet.cell(row=2, column=1)

print(type(cell))

print(cell.value)
'''
'''
pprint.pprint(sheet['A2:C4'])


g = sheet.iter_rows(min_row=2, max_row=4, min_col=1, max_col=3)

pprint.pprint(list(g))
'''

for u in range(6):
    for t in range(192):
        ruct = sheet.cell(row=t+1, column=u+2).value
        tDonw = (t+1)+1
        ructDown = sheet.cell(row=tDonw, column=u+2).value

        #if ruct == 0, 5, -5, 15:
        while ructDown == 204:
            tDonw += 1
            ructDown = sheet.cell(row=tDonw, column=u+2).value

        if ructDown == ruct:
            sheet.cell(row=(t+1)+1, column=u+2).value = ruct
        elif ruct == 15 and ructDown == -15:
            sheet.cell(row=(t+1)+1, column=u+2).value = 15
        

wb.save('./pyunyo.xlsx')

'''
r1c1 = ws.cell(row=2, column=3).value


def get_value_list(t_2d):
    return ([[cell.value for cell in row] for row in t_2d])

l_2d = get_value_list(sheet['B3:B112'])

pprint.pprint(l_2d, width=40)


print(wb.sheetnames)
'''


g_all = sheet.values

pprint.pprint(list(g_all), width=40)
